#-*- coding: utf-8 -*-
from openpyxl import load_workbook


# Load workbook.
wb = load_workbook(filename = 'pyfunc\\tpl\\ecortdata.xlsx')
ws = wb['Sheet1']
ws['A3'] = 'HAHA'
wb.save('01.xlsx')
wb.close()
print('done!')